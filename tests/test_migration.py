import shutil
import tempfile
import unittest
from copy import deepcopy
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from src.warranty_application_archive.approval_review import (
    ApprovalReviewRepository,
    apply_review_decisions,
    build_approval_review,
    export_approval_review_excel,
    import_excel_decisions,
)
from src.warranty_application_archive.constants import (
    APPROVAL_REVIEW_DATA_FILE_NAME,
    APPROVAL_REVIEW_EXCEL_FILE_NAME,
    DATA_FILE_NAME,
    TEMPLATE_FILE_NAME,
)
from src.warranty_application_archive.excel_export import export_excel
from src.warranty_application_archive.file_utils import sha256_file
from src.warranty_application_archive.migration import (
    apply_migration_plan,
    build_migration_plan,
    file_record,
    verify_backup,
)
from src.warranty_application_archive.repository import JsonRepository
from src.warranty_application_archive.workflows import (
    intake_applications,
    ingest_approval_pdfs,
    ingest_worker_lists,
)


PARSED_APPLICATION = {
    "项目名称": "测试项目",
    "质保单位": "测试单位",
    "分包单位": "",
    "质保负责人": "负责人",
    "质保负责人联系电话": "13800000000",
    "施工区域": "冷却塔",
    "施工开始时间": "2026-07-24",
    "施工结束时间": "2026-07-24",
    "时长天": 1,
    "施工内容": "维修冷塔",
    "施工负责人": "施工负责人",
    "施工负责人联系电话": "13900000000",
    "影响改动消防设备设施": "否",
    "影响堵塞应急疏散通道": "否",
    "危险作业": "",
}


class MigrationTest(unittest.TestCase):
    def _migrated_fixture(
        self, base: Path
    ) -> tuple[Path, dict[str, object], str]:
        primary = base / "质保作业申请单"
        primary.mkdir()
        stem = "2026-07-24_维修冷塔_质保作业申请单"
        (primary / f"{stem}.docx").write_bytes(b"word")
        (primary / f"{stem}.jpg").write_bytes(b"signed")
        (primary / f"{stem}_工人名单.jpg").write_bytes(b"workers")
        (primary / TEMPLATE_FILE_NAME).write_bytes(b"agreement")
        with patch(
            "src.warranty_application_archive.migration.legacy.parse_document",
            return_value=PARSED_APPLICATION,
        ):
            plan = build_migration_plan(primary)
        return primary, apply_migration_plan(plan), stem

    def test_migration_builds_case_directory_and_json(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            base = Path(temporary_dir)
            primary = base / "质保作业申请单"
            backup = base / "质保作业申请单_backup"
            primary.mkdir()
            stem = "2026-07-24_维修冷塔_质保作业申请单"
            (primary / f"{stem}.docx").write_bytes(b"word")
            (primary / f"{stem}.jpg").write_bytes(b"signed")
            (primary / f"{stem}_工人名单.jpg").write_bytes(b"workers")
            (primary / TEMPLATE_FILE_NAME).write_bytes(b"agreement")
            shutil.copytree(primary, backup)

            verify_backup(primary, backup)
            with patch(
                "src.warranty_application_archive.migration.legacy.parse_document",
                return_value=PARSED_APPLICATION,
            ):
                plan = build_migration_plan(primary)

            self.assertEqual(len(plan.applications), 1)
            self.assertEqual(plan.applications[0]["status"], "materials_ready")
            self.assertEqual(plan.applications[0]["missing_material_types"], [])

            dataset = apply_migration_plan(plan)
            repository = JsonRepository(primary)
            repository.save(dataset)
            excel_path = export_excel(dataset, primary)

            case_dir = primary / "_cases" / stem
            self.assertTrue((case_dir / f"{stem}.docx").is_file())
            self.assertTrue(
                (case_dir / f"{stem}_手签_01.jpg").is_file()
            )
            self.assertTrue(
                (case_dir / f"{stem}_施工人员名单_01.jpg").is_file()
            )
            self.assertTrue(
                (
                    case_dir
                    / "2026-07-24_维修冷塔_01 安全生产及消防安全协议（建工）.pdf"
                ).is_file()
            )
            self.assertTrue((primary / DATA_FILE_NAME).is_file())
            self.assertTrue(excel_path.is_file())
            workbook = load_workbook(excel_path, read_only=False, data_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "申请汇总",
                        "待补材料",
                        "待审批PDF",
                        "已完成",
                        "本次变更",
                        "说明",
                    ],
                )
                summary = workbook["申请汇总"]
                self.assertEqual(summary["A2"].value, "材料齐全，待审批PDF")
                self.assertEqual(summary["G2"].value, "维修冷塔")
                self.assertIsNotNone(summary["J2"].hyperlink)
            finally:
                workbook.close()

    def test_backup_mismatch_blocks_migration(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            base = Path(temporary_dir)
            primary = base / "primary"
            backup = base / "backup"
            primary.mkdir()
            backup.mkdir()
            (primary / "file.txt").write_text("primary", encoding="utf-8")
            (backup / "file.txt").write_text("changed", encoding="utf-8")

            with self.assertRaises(RuntimeError):
                verify_backup(primary, backup)

    def test_worker_list_and_approval_pdf_subworkflows(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            base = Path(temporary_dir)
            primary, dataset, stem = self._migrated_fixture(base)
            inbox = primary / "_inbox"
            inbox.mkdir(exist_ok=True)

            worker_source = inbox / "现场施工人员名单.jpg"
            worker_source.write_bytes(b"new workers")
            application = dataset["applications"][0]
            before_workers = len(
                application["materials"]["worker_list"]
            )
            worker_count = ingest_worker_lists(dataset, primary)
            self.assertEqual(worker_count, 1)
            self.assertEqual(
                len(application["materials"]["worker_list"]),
                before_workers + 1,
            )
            self.assertFalse(worker_source.exists())

            approval_source = (
                inbox / "工程类-主体质保施工_编号：202607240001.pdf"
            )
            approval_source.write_bytes(b"approval")
            dataset["unmatched_files"].append(
                file_record(
                    approval_source,
                    approval_source,
                    primary,
                    "approval_pdf",
                )
            )
            recognized_text = (
                "工程类-主体质保施工 申请编号：202607240001 "
                "施工区域：冷却塔 施工内容：维修冷塔 "
                "施工开始时间：2026年7月24日 "
                "施工结束时间：2026年7月24日"
            )
            with patch(
                "src.warranty_application_archive.workflows."
                "RecognitionService.pdf_text",
                return_value=recognized_text,
            ):
                approval_count = ingest_approval_pdfs(
                    dataset,
                    primary,
                    Path(__file__).resolve().parents[1],
                )

            self.assertEqual(approval_count, 1)
            target = (
                primary
                / "_cases"
                / stem
                / "工程类-主体质保施工_编号：202607240001.pdf"
            )
            self.assertTrue(target.is_file())
            self.assertEqual(application["status"], "approved")
            self.assertEqual(
                application["approval"]["application_no"],
                "202607240001",
            )

    def test_new_application_intake_creates_case(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            base = Path(temporary_dir)
            primary, dataset, _stem = self._migrated_fixture(base)
            inbox = primary / "_inbox"
            inbox.mkdir(exist_ok=True)
            incoming_word = inbox / "待处理申请.docx"
            incoming_word.write_bytes(b"new word")
            incoming_image = inbox / "待处理申请.jpg"
            incoming_image.write_bytes(b"new signed")
            parsed = {
                **PARSED_APPLICATION,
                "施工开始时间": "2026-07-25",
                "施工结束时间": "2026-07-25",
                "施工内容": "保温修复",
            }

            with patch(
                "src.warranty_application_archive.workflows."
                "legacy.parse_document",
                return_value=parsed,
            ):
                count = intake_applications(
                    dataset,
                    primary,
                    Path(__file__).resolve().parents[1],
                )

            self.assertEqual(count, 1)
            application = dataset["applications"][-1]
            self.assertEqual(
                application["case_name"],
                "2026-07-25_保温修复_质保作业申请单",
            )
            case_dir = (
                primary
                / "_cases"
                / "2026-07-25_保温修复_质保作业申请单"
            )
            self.assertTrue(
                (
                    case_dir
                    / "2026-07-25_保温修复_质保作业申请单.docx"
                ).is_file()
            )
            self.assertTrue(
                (
                    case_dir
                    / "2026-07-25_保温修复_质保作业申请单_手签_01.jpg"
                ).is_file()
            )

    def test_human_approval_review_updates_formal_data(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            base = Path(temporary_dir)
            primary, dataset, stem = self._migrated_fixture(base)
            inbox = primary / "_inbox"
            inbox.mkdir(exist_ok=True)
            source = (
                inbox / "工程类-主体质保施工_编号：202607240099.pdf"
            )
            source.write_bytes(b"approval requiring human review")
            source_record = file_record(
                source,
                source,
                primary,
                "approval_pdf",
            )
            dataset["unmatched_files"].append(source_record)
            dataset.setdefault("recognition_cache", {})[
                sha256_file(source)
            ] = {
                "text": (
                    "工程类主体质保施工申请编号202607240099"
                    "施工区域冷却塔施工开始时间施工结束时间"
                    "2026~07~23至2026~07~25时长3天"
                    "施工内容维修冷塔施工负责人测试人员"
                ),
                "method": "test",
            }
            outside_inbox = primary / "历史未匹配审批.pdf"
            outside_inbox.write_bytes(b"outside inbox")
            outside_record = file_record(
                outside_inbox,
                outside_inbox,
                primary,
                "approval_pdf",
            )
            dataset["unmatched_files"].append(outside_record)
            dataset["recognition_cache"][sha256_file(outside_inbox)] = {
                "text": "施工区域冷却塔施工内容维修冷塔",
                "method": "test",
            }

            review = build_approval_review(
                dataset,
                primary,
                Path(__file__).resolve().parents[1],
            )
            self.assertEqual(len(review["pending_reviews"]), 1)
            candidate = review["pending_reviews"][0]
            self.assertEqual(candidate["case"]["case_name"], stem)
            self.assertEqual(candidate["pdf"]["start"], "2026-07-23")
            self.assertGreaterEqual(candidate["confidence"], 80)
            self.assertEqual(review["unresolved_pdfs"], [])

            excel_path = export_approval_review_excel(review, primary)
            workbook = load_workbook(excel_path)
            try:
                sheet = workbook["待审核"]
                sheet["A2"] = "确认匹配"
                sheet["C2"] = "人工核对施工内容与日期后确认"
                workbook.save(excel_path)
            finally:
                workbook.close()

            decisions = import_excel_decisions(review, primary)
            applied = apply_review_decisions(
                dataset,
                review,
                decisions,
                primary,
            )
            self.assertEqual(applied, 1)
            application = dataset["applications"][0]
            self.assertEqual(application["status"], "approved")
            self.assertEqual(
                application["approval"]["match_source"],
                "human_review",
            )
            self.assertEqual(len(dataset["unmatched_files"]), 1)
            self.assertEqual(
                dataset["unmatched_files"][0]["path"],
                "历史未匹配审批.pdf",
            )
            self.assertFalse(source.exists())
            self.assertTrue(
                (
                    primary
                    / "_cases"
                    / stem
                    / "工程类-主体质保施工_编号：202607240099.pdf"
                ).is_file()
            )
            self.assertTrue(
                (primary / APPROVAL_REVIEW_EXCEL_FILE_NAME).is_file()
            )

    def test_review_outputs_only_top_strict_candidate_per_inbox_pdf(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            base = Path(temporary_dir)
            primary, dataset, _stem = self._migrated_fixture(base)
            inbox = primary / "_inbox"
            inbox.mkdir(exist_ok=True)

            first_application = dataset["applications"][0]
            first_application["application"]["施工开始时间"] = "2026-01-01"
            first_application["application"]["施工结束时间"] = "2026-01-02"
            second_application = deepcopy(first_application)
            second_application["case_id"] = "second-case"
            second_application["case_name"] = (
                "2026-08-18_维修冷塔_质保作业申请单"
            )
            second_application["case_directory"] = (
                "_cases/2026-08-18_维修冷塔_质保作业申请单"
            )
            second_application["application"]["施工开始时间"] = "2026-08-18"
            second_application["application"]["施工结束时间"] = "2026-08-20"
            (
                primary / second_application["case_directory"]
            ).mkdir(parents=True)
            dataset["applications"].append(second_application)

            matching_pdf = inbox / "工程类-主体质保施工_编号：202608190001.pdf"
            matching_pdf.write_bytes(b"top candidate")
            dataset["unmatched_files"].append(
                file_record(
                    matching_pdf,
                    matching_pdf,
                    primary,
                    "approval_pdf",
                )
            )
            dataset["unmatched_files"].append(
                deepcopy(dataset["unmatched_files"][-1])
            )
            matching_hash = sha256_file(matching_pdf)
            dataset.setdefault("recognition_cache", {})[matching_hash] = {
                "text": (
                    "工程类主体质保施工申请编号202608190001"
                    "施工区域冷却塔施工开始时间2026年8月19日"
                    "施工结束时间2026年8月20日"
                    "施工内容维修冷塔施工负责人测试人员"
                ),
                "method": "ocr",
            }

            unresolved_pdf = (
                inbox / "工程类-主体质保施工_编号：202608190002.pdf"
            )
            unresolved_pdf.write_bytes(b"no strict candidate")
            dataset["unmatched_files"].append(
                file_record(
                    unresolved_pdf,
                    unresolved_pdf,
                    primary,
                    "approval_pdf",
                )
            )
            unresolved_hash = sha256_file(unresolved_pdf)
            dataset["recognition_cache"][unresolved_hash] = {
                "text": (
                    "工程类主体质保施工申请编号202608190002"
                    "施工区域冷却塔施工开始时间2026年7月24日"
                    "施工结束时间2026年7月24日"
                    "施工内容更换UV灯管施工负责人测试人员"
                ),
                "method": "plain",
            }

            nested = inbox / "nested"
            nested.mkdir()
            nested_pdf = nested / "嵌套目录审批.pdf"
            nested_pdf.write_bytes(b"nested")
            dataset["unmatched_files"].append(
                file_record(
                    nested_pdf,
                    nested_pdf,
                    primary,
                    "approval_pdf",
                )
            )
            dataset["recognition_cache"][sha256_file(nested_pdf)] = {
                "text": "施工区域冷却塔施工内容维修冷塔",
                "method": "test",
            }

            review = build_approval_review(
                dataset,
                primary,
                Path(__file__).resolve().parents[1],
            )

            self.assertEqual(len(review["pending_reviews"]), 1)
            candidate = review["pending_reviews"][0]
            self.assertEqual(candidate["pdf"]["sha256"], matching_hash)
            self.assertEqual(
                candidate["case"]["case_id"],
                "second-case",
            )
            self.assertEqual(candidate["strict_candidate_count"], 2)
            self.assertEqual(candidate["confidence"], 98)
            self.assertEqual(candidate["runner_up_confidence"], 80)
            self.assertEqual(candidate["confidence_gap"], 18)
            self.assertEqual(len(review["unresolved_pdfs"]), 1)
            self.assertEqual(
                review["unresolved_pdfs"][0]["pdf"]["sha256"],
                unresolved_hash,
            )
            self.assertIn(
                "施工内容",
                review["unresolved_pdfs"][0]["reason"],
            )

            review_path = ApprovalReviewRepository(primary).save(review)
            self.assertEqual(
                review_path.name,
                APPROVAL_REVIEW_DATA_FILE_NAME,
            )
            excel_path = export_approval_review_excel(review, primary)
            self.assertEqual(
                excel_path.name,
                APPROVAL_REVIEW_EXCEL_FILE_NAME,
            )
            workbook = load_workbook(excel_path, data_only=False)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["待审核", "无严格候选", "已处理决定", "说明"],
                )
                pending_sheet = workbook["待审核"]
                self.assertEqual(pending_sheet.max_row, 2)
                self.assertEqual(pending_sheet["D2"].value, 0.98)
                self.assertIn("待人工审核", str(pending_sheet["B2"].value))
                unresolved_sheet = workbook["无严格候选"]
                self.assertEqual(unresolved_sheet.max_row, 2)
                self.assertEqual(
                    unresolved_sheet["C2"].value,
                    unresolved_pdf.name,
                )
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
