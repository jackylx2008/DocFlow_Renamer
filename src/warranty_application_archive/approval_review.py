from __future__ import annotations

import json
import os
import re
import shutil
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import legacy
from .constants import (
    APPROVAL_PDF_ROLE,
    APPROVAL_REVIEW_DATA_FILE_NAME,
    APPROVAL_REVIEW_EXCEL_FILE_NAME,
    APPROVAL_REVIEW_SCHEMA_VERSION,
    INBOX_DIR_NAME,
    INTERNAL_DIR_NAME,
    LEGACY_APPROVAL_REVIEW_DATA_FILE_NAME,
    LEGACY_APPROVAL_REVIEW_EXCEL_FILE_NAME,
    LEGACY_DIR_NAME,
)
from .file_utils import atomic_replace_text, ensure_within, sha256_file
from .migration import CASE_NAMESPACE
from .recognition import RecognitionService
from .workflows import archive_reviewed_approval_pdf


PENDING_HEADERS = [
    "审核结果",
    "审核处理状态",
    "人工备注",
    "匹配置信度",
    "置信度等级",
    "严格候选数",
    "次高置信度",
    "领先差值",
    "匹配依据",
    "识别方式",
    "审批PDF文件",
    "审批编号",
    "审批施工区域",
    "审批施工开始",
    "审批施工结束",
    "审批施工内容",
    "候选案卷",
    "案卷状态",
    "案卷施工区域",
    "案卷施工开始",
    "案卷施工结束",
    "案卷施工内容",
    "PDF链接",
    "案卷目录",
    "审核ID",
    "PDF SHA-256",
    "案卷ID",
]
UNRESOLVED_HEADERS = [
    "处理状态",
    "未形成候选原因",
    "审批PDF文件",
    "审批编号",
    "审批施工区域",
    "审批施工开始",
    "审批施工结束",
    "审批施工内容",
    "识别方式",
    "PDF链接",
    "PDF SHA-256",
    "PDF相对路径",
]
DECISION_OPTIONS = ("待审核", "确认匹配", "排除")
MATCH_RULE_VERSION = "strict-non-date-relaxed-date-v1"
NON_DATE_CONFIDENCE = 80
CASE_STATUS_LABELS = {
    "materials_incomplete": "材料待补充",
    "materials_ready": "材料齐全，待审批PDF",
    "approval_pdf_unmatched": "审批PDF待确认",
    "approved": "审批完成",
    "needs_review": "待人工确认",
}
CASE_NAME_RE = re.compile(
    r"^(?P<date>\d{4}-\d{2}-\d{2})_(?P<content>.+?)_质保作业申请单"
)


def _now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _empty_review(root: Path) -> dict[str, Any]:
    return {
        "schema_version": APPROVAL_REVIEW_SCHEMA_VERSION,
        "data_root": str(root.resolve()),
        "source_dataset_revision": 0,
        "generated_at": _now(),
        "match_rule_version": MATCH_RULE_VERSION,
        "pending_reviews": [],
        "unresolved_pdfs": [],
        "decisions": [],
    }


class ApprovalReviewRepository:
    def __init__(self, root: Path) -> None:
        self.root = root.resolve()
        self.path = self.root / APPROVAL_REVIEW_DATA_FILE_NAME
        self.legacy_path = (
            self.root / LEGACY_APPROVAL_REVIEW_DATA_FILE_NAME
        )

    def load(self) -> dict[str, Any]:
        source = self.path if self.path.is_file() else self.legacy_path
        if not source.is_file():
            return _empty_review(self.root)
        data = json.loads(source.read_text(encoding="utf-8"))
        schema_version = int(data.get("schema_version") or 0)
        if schema_version not in {1, APPROVAL_REVIEW_SCHEMA_VERSION}:
            raise ValueError(
                "不支持的审批 PDF 审核数据版本: "
                f"{data.get('schema_version')}"
            )
        data["schema_version"] = APPROVAL_REVIEW_SCHEMA_VERSION
        return data

    def save(self, data: dict[str, Any]) -> Path:
        data["schema_version"] = APPROVAL_REVIEW_SCHEMA_VERSION
        data["data_root"] = str(self.root)
        data["updated_at"] = _now()
        atomic_replace_text(
            self.path,
            json.dumps(data, ensure_ascii=False, indent=2),
        )
        _archive_legacy_artifact(self.root, self.legacy_path)
        return self.path


def _archive_legacy_artifact(root: Path, path: Path) -> None:
    if not path.is_file():
        return
    archive_dir = root / INTERNAL_DIR_NAME / LEGACY_DIR_NAME
    archive_dir.mkdir(parents=True, exist_ok=True)
    target = archive_dir / path.name
    if target.exists():
        target = archive_dir / (
            f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            f"{path.suffix}"
        )
    shutil.move(str(path), str(target))


def _date_from_text(value: str) -> str:
    return legacy.normalize_date_for_pdf_match(value)


def _case_values(application: dict[str, Any]) -> dict[str, str]:
    business = application.get("application") or {}
    case_name = str(application.get("case_name") or "")
    name_match = CASE_NAME_RE.match(case_name)
    return {
        "area": str(business.get("施工区域") or ""),
        "start": (
            _date_from_text(str(business.get("施工开始时间") or ""))
            or (name_match.group("date") if name_match else "")
        ),
        "end": _date_from_text(str(business.get("施工结束时间") or "")),
        "content": (
            str(business.get("施工内容") or "")
            or (name_match.group("content") if name_match else "")
        ),
    }


def _between(text: str, start: str, ends: tuple[str, ...]) -> str:
    start_index = text.find(start)
    if start_index < 0:
        return ""
    value_start = start_index + len(start)
    value_end = len(text)
    for end in ends:
        index = text.find(end, value_start)
        if index >= 0:
            value_end = min(value_end, index)
    return text[value_start:value_end].strip("~:：-—")


def _pdf_values(text: str) -> dict[str, str]:
    normalized = legacy.normalize_match_text(text)
    construction_segment = _between(
        normalized,
        "施工开始时间",
        ("时长", "施工内容"),
    )
    construction_dates = [
        legacy.date_match_groups_to_iso(match.groups())
        for match in legacy.PDF_DATE_TOKEN_RE.finditer(construction_segment)
    ]
    construction_dates = [value for value in construction_dates if value]
    start = construction_dates[0] if construction_dates else ""
    end = construction_dates[1] if len(construction_dates) > 1 else ""
    if not start:
        match = legacy.PDF_START_DATE_RE.search(normalized)
        start = (
            legacy.date_match_groups_to_iso(match.groups()) if match else ""
        )
    if not end:
        match = legacy.PDF_END_DATE_RE.search(normalized)
        end = legacy.date_match_groups_to_iso(match.groups()) if match else ""
    return {
        "area": _between(
            normalized,
            "施工区域",
            ("施工开始时间", "施工开始日期"),
        ),
        "start": start,
        "end": end,
        "content": _between(
            normalized,
            "施工内容",
            (
                "施工负责人",
                "影响改动消防",
                "影响堵塞",
                "危险作业",
                "附件",
                "审批记录",
            ),
        ),
    }


def _date_delta(first: str, second: str) -> int | None:
    try:
        return abs((date.fromisoformat(first) - date.fromisoformat(second)).days)
    except (TypeError, ValueError):
        return None


def _strict_non_date_match(
    pdf_text: str,
    case_values: dict[str, str],
) -> tuple[bool, bool]:
    normalized_pdf = legacy.normalize_match_text(pdf_text)
    area_key = legacy.normalize_match_text(case_values["area"])
    content_key = legacy.normalize_match_text(case_values["content"])
    if not area_key or not content_key:
        return False, False

    content_keys = [content_key]
    if content_key.startswith(area_key):
        shortened = content_key[len(area_key) :]
        if shortened:
            content_keys.append(shortened)
    return (
        area_key in normalized_pdf,
        any(key in normalized_pdf for key in content_keys),
    )


def _date_points(delta: int | None, maximum: int) -> int:
    if delta is None:
        return 0
    if delta == 0:
        return maximum
    if delta <= 3:
        return maximum - 2
    if delta <= 7:
        return max(0, maximum - 4)
    if delta <= 14:
        return max(0, maximum - 7)
    if delta <= 31:
        return max(0, maximum - 10)
    return 0


def _candidate_confidence(
    pdf_values: dict[str, str],
    case_values: dict[str, str],
) -> tuple[int, list[str]]:
    evidence: list[str] = []
    start_delta = _date_delta(case_values["start"], pdf_values["start"])
    end_delta = _date_delta(case_values["end"], pdf_values["end"])
    confidence = (
        NON_DATE_CONFIDENCE
        + _date_points(start_delta, 12)
        + _date_points(end_delta, 8)
    )
    evidence.extend(["施工内容严格命中", "施工区域严格命中"])
    if start_delta is None:
        evidence.append("开始日期缺失，不参与淘汰")
    else:
        evidence.append(f"开始日期相差{start_delta}天")
    if end_delta is None:
        evidence.append("结束日期缺失，不参与淘汰")
    else:
        evidence.append(f"结束日期相差{end_delta}天")
    return min(100, confidence), evidence


def _confidence_level(confidence: int) -> str:
    if confidence >= 95:
        return "高"
    if confidence >= 85:
        return "中"
    return "低"


def _selection_confidence(
    candidate_confidence: int,
    runner_up_confidence: int | None,
) -> int:
    if runner_up_confidence is None:
        return candidate_confidence
    gap = max(0, candidate_confidence - runner_up_confidence)
    ambiguity_limit = 60 + min(40, gap * 3)
    return min(candidate_confidence, ambiguity_limit)


def _review_id(pdf_hash: str, case_id: str) -> str:
    return str(
        uuid.uuid5(
            CASE_NAMESPACE,
            f"approval-review:{pdf_hash}:{case_id}",
        )
    )


def _direct_inbox_pdf(root: Path, item: dict[str, Any]) -> Path | None:
    try:
        path = ensure_within(
            root / Path(str(item.get("path") or "")),
            root,
        )
    except ValueError:
        return None
    inbox = (root / INBOX_DIR_NAME).resolve()
    if path.parent != inbox or path.suffix.lower() != ".pdf":
        return None
    return path


def build_approval_review(
    dataset: dict[str, Any],
    root: Path,
    repo_root: Path,
    existing: dict[str, Any] | None = None,
) -> dict[str, Any]:
    root = root.resolve()
    previous = existing or _empty_review(root)
    decisions = list(previous.get("decisions") or [])
    excluded_ids = {
        str(item.get("review_id") or "")
        for item in decisions
        if item.get("decision") == "排除"
    }
    applications = [
        application
        for application in dataset.get("applications") or []
        if not (application.get("approval") or {}).get("pdfs")
    ]
    pending: list[dict[str, Any]] = []
    unresolved: list[dict[str, Any]] = []
    seen_pdf_hashes: set[str] = set()
    with RecognitionService(dataset, repo_root) as recognition:
        for pdf_item in dataset.get("unmatched_files") or []:
            if pdf_item.get("role") != APPROVAL_PDF_ROLE:
                continue
            pdf_path = _direct_inbox_pdf(root, pdf_item)
            if pdf_path is None or not pdf_path.is_file():
                continue
            pdf_hash = sha256_file(pdf_path)
            if pdf_hash in seen_pdf_hashes:
                continue
            seen_pdf_hashes.add(pdf_hash)
            text = recognition.pdf_text(pdf_path)
            values = _pdf_values(text)
            application_no = (
                legacy.extract_pdf_application_no_from_name(pdf_path.name)
                or legacy.extract_pdf_rename_application_no(text)
            )
            recognition_method = str(
                (
                    (dataset.get("recognition_cache") or {})
                    .get(pdf_hash, {})
                    .get("method", "unknown")
                )
            )
            pdf_payload = {
                "path": str(pdf_item.get("path") or ""),
                "file_name": pdf_path.name,
                "sha256": pdf_hash,
                "application_no": application_no,
                "recognition_method": recognition_method,
                **values,
            }
            ranked: list[
                tuple[
                    int,
                    str,
                    dict[str, Any],
                    dict[str, str],
                    list[str],
                    str,
                ]
            ] = []
            any_area_match = False
            any_content_match = False
            strict_matches = 0
            for application in applications:
                case_values = _case_values(application)
                area_match, content_match = _strict_non_date_match(
                    text,
                    case_values,
                )
                any_area_match = any_area_match or area_match
                any_content_match = any_content_match or content_match
                if not (area_match and content_match):
                    continue
                strict_matches += 1
                confidence, evidence = _candidate_confidence(
                    values,
                    case_values,
                )
                case_id = str(application.get("case_id") or "")
                review_id = _review_id(pdf_hash, case_id)
                if review_id in excluded_ids:
                    continue
                ranked.append(
                    (
                        confidence,
                        str(application.get("case_name") or ""),
                        application,
                        case_values,
                        evidence,
                        review_id,
                    )
                )
            ranked.sort(key=lambda item: (-item[0], item[1]))
            if not ranked:
                if strict_matches:
                    reason = "所有严格候选均已被人工排除"
                elif not applications:
                    reason = "没有可匹配的未审批案卷"
                elif not any_content_match:
                    reason = "施工内容没有严格文本命中"
                elif not any_area_match:
                    reason = "施工区域没有严格文本命中"
                else:
                    reason = "施工区域和施工内容未在同一案卷同时严格命中"
                unresolved.append(
                    {
                        "status": "无严格候选",
                        "reason": reason,
                        "pdf": pdf_payload,
                    }
                )
                continue

            (
                candidate_confidence,
                _case_name,
                application,
                case_values,
                evidence,
                review_id,
            ) = ranked[0]
            runner_up_confidence = ranked[1][0] if len(ranked) > 1 else None
            confidence_gap = (
                candidate_confidence - runner_up_confidence
                if runner_up_confidence is not None
                else None
            )
            confidence = _selection_confidence(
                candidate_confidence,
                runner_up_confidence,
            )
            case_id = str(application.get("case_id") or "")
            ambiguity_evidence = (
                [
                    f"候选自身匹配度{candidate_confidence}%",
                    f"次高匹配度{runner_up_confidence}%",
                    f"领先{confidence_gap}个百分点",
                    f"综合选择置信度{confidence}%",
                ]
                if runner_up_confidence is not None
                else [f"唯一严格候选，置信度{confidence}%"]
            )
            pending.append(
                {
                    "review_id": review_id,
                    "decision": "待审核",
                    "review_note": "",
                    "confidence": confidence,
                    "candidate_match_confidence": candidate_confidence,
                    "confidence_level": _confidence_level(confidence),
                    "strict_candidate_count": len(ranked),
                    "runner_up_confidence": runner_up_confidence,
                    "confidence_gap": confidence_gap,
                    "matching_evidence": "；".join(
                        [
                            *evidence,
                            *ambiguity_evidence,
                            f"严格候选{len(ranked)}个，仅展示最高项",
                        ]
                    ),
                    "pdf": pdf_payload,
                    "case": {
                        "case_id": case_id,
                        "case_name": str(application.get("case_name") or ""),
                        "case_directory": str(
                            application.get("case_directory") or ""
                        ),
                        "status": str(application.get("status") or ""),
                        **case_values,
                    },
                }
            )
    return {
        "schema_version": APPROVAL_REVIEW_SCHEMA_VERSION,
        "data_root": str(root),
        "source_dataset_revision": int(
            dataset.get("dataset_revision") or 0
        ),
        "generated_at": _now(),
        "match_rule_version": MATCH_RULE_VERSION,
        "pending_reviews": pending,
        "unresolved_pdfs": unresolved,
        "decisions": decisions,
    }


def _pending_row(item: dict[str, Any]) -> list[Any]:
    pdf = item.get("pdf") or {}
    case = item.get("case") or {}
    runner_up = item.get("runner_up_confidence")
    confidence_gap = item.get("confidence_gap")
    return [
        item.get("decision", "待审核"),
        "待人工审核",
        item.get("review_note", ""),
        float(item.get("confidence", 0)) / 100,
        item.get("confidence_level", ""),
        item.get("strict_candidate_count", 0),
        (
            float(runner_up) / 100
            if isinstance(runner_up, (int, float))
            else ""
        ),
        (
            float(confidence_gap) / 100
            if isinstance(confidence_gap, (int, float))
            else ""
        ),
        item.get("matching_evidence", ""),
        pdf.get("recognition_method", ""),
        pdf.get("file_name", ""),
        pdf.get("application_no", ""),
        pdf.get("area", ""),
        pdf.get("start", ""),
        pdf.get("end", ""),
        pdf.get("content", ""),
        case.get("case_name", ""),
        CASE_STATUS_LABELS.get(
            str(case.get("status") or ""),
            case.get("status", ""),
        ),
        case.get("area", ""),
        case.get("start", ""),
        case.get("end", ""),
        case.get("content", ""),
        "打开审批PDF",
        "打开案卷目录",
        item.get("review_id", ""),
        pdf.get("sha256", ""),
        case.get("case_id", ""),
    ]


def _unresolved_row(item: dict[str, Any]) -> list[Any]:
    pdf = item.get("pdf") or {}
    return [
        item.get("status", "无严格候选"),
        item.get("reason", ""),
        pdf.get("file_name", ""),
        pdf.get("application_no", ""),
        pdf.get("area", ""),
        pdf.get("start", ""),
        pdf.get("end", ""),
        pdf.get("content", ""),
        pdf.get("recognition_method", ""),
        "打开审批PDF",
        pdf.get("sha256", ""),
        pdf.get("path", ""),
    ]


def _format_table_sheet(sheet: Any, rows: int, columns: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(columns)}{max(2, rows + 1)}"
    )
    header_fill = PatternFill("solid", fgColor="44546A")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 28


def export_approval_review_excel(
    review: dict[str, Any], root: Path
) -> Path:
    root = root.resolve()
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    sheet = workbook.active
    sheet.title = "待审核"
    sheet.append(PENDING_HEADERS)
    pending = list(review.get("pending_reviews") or [])
    for item in pending:
        sheet.append(_pending_row(item))
    _format_table_sheet(sheet, len(pending), len(PENDING_HEADERS))
    if pending:
        table = Table(
            displayName="ApprovalReviewCandidates",
            ref=(
                f"A1:{get_column_letter(len(PENDING_HEADERS))}"
                f"{len(pending) + 1}"
            ),
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        validation = DataValidation(
            type="list",
            formula1='"待审核,确认匹配,排除"',
            allow_blank=False,
        )
        validation.error = "请选择：待审核、确认匹配或排除"
        validation.errorTitle = "审核结果无效"
        sheet.add_data_validation(validation)
        validation.add(f"A2:A{len(pending) + 1}")
        sheet.conditional_formatting.add(
            f"A2:A{len(pending) + 1}",
            FormulaRule(
                formula=['$A2="确认匹配"'],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        sheet.conditional_formatting.add(
            f"A2:A{len(pending) + 1}",
            FormulaRule(
                formula=['$A2="排除"'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        sheet.conditional_formatting.add(
            f"D2:D{len(pending) + 1}",
            FormulaRule(
                formula=["$D2>=0.95"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        sheet.conditional_formatting.add(
            f"D2:D{len(pending) + 1}",
            FormulaRule(
                formula=['AND($D2>=0.85,$D2<0.95)'],
                fill=PatternFill("solid", fgColor="FFF2CC"),
            ),
        )
        sheet.conditional_formatting.add(
            f"D2:D{len(pending) + 1}",
            FormulaRule(
                formula=["$D2<0.85"],
                fill=PatternFill("solid", fgColor="FCE4D6"),
            ),
        )
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    for row_index in range(2, len(pending) + 2):
        sheet.cell(row=row_index, column=2).value = (
            f'=IF(A{row_index}="确认匹配","已确认匹配，待执行回写",'
            f'IF(A{row_index}="排除","已排除，待执行回写","待人工审核"))'
        )
        for column_index in (1, 3):
            sheet.cell(row=row_index, column=column_index).fill = editable_fill
            sheet.cell(row=row_index, column=column_index).protection = Protection(
                locked=False
            )
        item = pending[row_index - 2]
        pdf_path = root / Path(str((item.get("pdf") or {}).get("path") or ""))
        case_path = root / Path(
            str((item.get("case") or {}).get("case_directory") or "")
        )
        sheet.cell(row=row_index, column=23).hyperlink = str(pdf_path.resolve())
        sheet.cell(row=row_index, column=24).hyperlink = str(case_path.resolve())
        for column_index in (23, 24):
            sheet.cell(row=row_index, column=column_index).font = Font(
                color="0563C1", underline="single"
            )
    widths = [
        12, 24, 28, 12, 10, 12, 12, 10, 42, 12, 38, 16, 22, 13,
        13, 32, 42, 24, 22, 13, 13, 32, 16, 16, 38, 68, 38,
    ]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(2, len(pending) + 2):
        for column_index in (4, 7, 8):
            sheet.cell(row=row_index, column=column_index).number_format = "0%"
    for column_index in (25, 26, 27):
        sheet.column_dimensions[get_column_letter(column_index)].hidden = True

    unresolved = workbook.create_sheet("无严格候选")
    unresolved.append(UNRESOLVED_HEADERS)
    unresolved_items = list(review.get("unresolved_pdfs") or [])
    for item in unresolved_items:
        unresolved.append(_unresolved_row(item))
    _format_table_sheet(
        unresolved,
        len(unresolved_items),
        len(UNRESOLVED_HEADERS),
    )
    if unresolved_items:
        unresolved_table = Table(
            displayName="ApprovalReviewUnresolved",
            ref=(
                f"A1:{get_column_letter(len(UNRESOLVED_HEADERS))}"
                f"{len(unresolved_items) + 1}"
            ),
        )
        unresolved_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
            showColumnStripes=False,
        )
        unresolved.add_table(unresolved_table)
    for row_index, item in enumerate(unresolved_items, start=2):
        pdf_path = root / Path(str((item.get("pdf") or {}).get("path") or ""))
        unresolved.cell(row=row_index, column=10).hyperlink = str(
            pdf_path.resolve()
        )
        unresolved.cell(row=row_index, column=10).font = Font(
            color="0563C1",
            underline="single",
        )
        unresolved.cell(row=row_index, column=1).fill = PatternFill(
            "solid",
            fgColor="FCE4D6",
        )
    unresolved_widths = (
        16, 38, 42, 16, 24, 13, 13, 32, 12, 16, 68, 48,
    )
    for index, width in enumerate(unresolved_widths, start=1):
        unresolved.column_dimensions[get_column_letter(index)].width = width
    for column_index in (11, 12):
        unresolved.column_dimensions[get_column_letter(column_index)].hidden = True

    history = workbook.create_sheet("已处理决定")
    history_headers = [
        "处理时间",
        "审核结果",
        "回写状态",
        "人工备注",
        "审批PDF文件",
        "PDF SHA-256",
        "候选案卷",
        "案卷ID",
        "审核ID",
    ]
    history.append(history_headers)
    for item in review.get("decisions") or []:
        history.append(
            [
                item.get("decided_at", ""),
                item.get("decision", ""),
                (
                    "已归档，正式案卷状态已更新为审批完成"
                    if item.get("decision") == "确认匹配"
                    else "已记录排除"
                ),
                item.get("review_note", ""),
                item.get("pdf_file_name", ""),
                item.get("pdf_sha256", ""),
                item.get("case_name", ""),
                item.get("case_id", ""),
                item.get("review_id", ""),
            ]
        )
    _format_table_sheet(
        history,
        len(review.get("decisions") or []),
        len(history_headers),
    )
    for column, width in zip(
        "ABCDEFGHI", (22, 12, 38, 32, 42, 68, 42, 38, 38)
    ):
        history.column_dimensions[column].width = width

    notes = workbook.create_sheet("说明")
    notes.sheet_view.showGridLines = False
    instructions = [
        ["审批 PDF 人工匹配审核", ""],
        [
            "1",
            "只处理 _inbox 第一层中尚未自动匹配的审批 PDF；"
            "每个 PDF 在“待审核”中只显示置信度最高的一个严格候选。",
        ],
        [
            "匹配规则",
            "施工区域和施工内容必须同时严格文本命中；日期不淘汰候选，"
            "只按相差天数降低置信度。",
        ],
        [
            "置信度",
            "非日期严格命中占 80%，开始和结束日期合计占 20%；"
            "“严格候选数、次高置信度、领先差值”用于判断第一名是否稳定。",
        ],
        [
            "无严格候选",
            "没有通过区域和内容严格门槛的 PDF 单独列在“无严格候选”，"
            "不会强行推荐案卷。",
        ],
        ["2", "只修改黄色的“审核结果”和“人工备注”列。"],
        [
            "3",
            "选择“排除”后重新生成审核表，会在剩余严格候选中选择下一名。",
        ],
        [
            "4",
            "保存并关闭 Excel 后，运行："
            "python warranty_application_archive.py apply-approval-review",
        ],
        ["5", "命令先保存审核决定，再更新正式 JSON，并从正式 JSON 重建正式汇总 Excel。"],
        ["正式数据版本", review.get("source_dataset_revision", 0)],
        ["生成时间", review.get("generated_at", "")],
        ["匹配规则版本", review.get("match_rule_version", "")],
        ["待审核 PDF", len(pending)],
        ["无严格候选 PDF", len(unresolved_items)],
    ]
    for row in instructions:
        notes.append(row)
    notes.merge_cells("A1:B1")
    notes["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="44546A")
    notes["A1"].alignment = Alignment(horizontal="center")
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 100
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    output = root / APPROVAL_REVIEW_EXCEL_FILE_NAME
    temporary = root / f".{APPROVAL_REVIEW_EXCEL_FILE_NAME}.tmp.xlsx"
    workbook.save(temporary)
    try:
        os.replace(temporary, output)
    except PermissionError as exc:
        temporary.unlink(missing_ok=True)
        raise PermissionError(
            f"审核 Excel 正在使用，请先关闭后重试: {output}"
        ) from exc
    _archive_legacy_artifact(
        root,
        root / LEGACY_APPROVAL_REVIEW_EXCEL_FILE_NAME,
    )
    return output


def import_excel_decisions(
    review: dict[str, Any], root: Path
) -> list[dict[str, Any]]:
    path = root.resolve() / APPROVAL_REVIEW_EXCEL_FILE_NAME
    if not path.is_file():
        legacy_path = (
            root.resolve() / LEGACY_APPROVAL_REVIEW_EXCEL_FILE_NAME
        )
        path = legacy_path if legacy_path.is_file() else path
    if not path.is_file():
        raise FileNotFoundError(f"待人工审核匹配 PDF Excel 不存在: {path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if "待审核" not in workbook.sheetnames:
            raise ValueError("审核 Excel 缺少“待审核”工作表")
        sheet = workbook["待审核"]
        headers = {
            str(cell.value or "").strip(): cell.column
            for cell in sheet[1]
        }
        missing = [header for header in PENDING_HEADERS if header not in headers]
        if missing:
            raise ValueError(f"审核 Excel 缺少列: {', '.join(missing)}")
        expected = {
            str(item.get("review_id") or ""): item
            for item in review.get("pending_reviews") or []
        }
        imported: list[dict[str, Any]] = []
        seen_ids: set[str] = set()
        for row_index in range(2, sheet.max_row + 1):
            review_id = str(
                sheet.cell(row_index, headers["审核ID"]).value or ""
            ).strip()
            if not review_id:
                continue
            if review_id in seen_ids:
                raise ValueError(f"审核 Excel 中审核ID重复: {review_id}")
            seen_ids.add(review_id)
            item = expected.get(review_id)
            if item is None:
                raise ValueError(
                    f"审核 Excel 含有过期或未知审核ID: {review_id}"
                )
            decision = str(
                sheet.cell(row_index, headers["审核结果"]).value or ""
            ).strip()
            if decision not in DECISION_OPTIONS:
                raise ValueError(
                    f"第 {row_index} 行审核结果无效: {decision}"
                )
            pdf_hash = str(
                sheet.cell(row_index, headers["PDF SHA-256"]).value or ""
            ).strip()
            case_id = str(
                sheet.cell(row_index, headers["案卷ID"]).value or ""
            ).strip()
            if pdf_hash != str((item.get("pdf") or {}).get("sha256") or ""):
                raise ValueError(f"第 {row_index} 行 PDF SHA-256 被修改")
            if case_id != str((item.get("case") or {}).get("case_id") or ""):
                raise ValueError(f"第 {row_index} 行案卷ID被修改")
            if decision != "待审核":
                imported.append(
                    {
                        **item,
                        "decision": decision,
                        "review_note": str(
                            sheet.cell(
                                row_index, headers["人工备注"]
                            ).value
                            or ""
                        ).strip(),
                    }
                )
        return imported
    finally:
        workbook.close()


def apply_review_decisions(
    dataset: dict[str, Any],
    review: dict[str, Any],
    decisions: list[dict[str, Any]],
    root: Path,
) -> int:
    confirmed = [
        item for item in decisions if item.get("decision") == "确认匹配"
    ]
    counts: dict[str, int] = {}
    selected_cases: set[str] = set()
    for item in confirmed:
        pdf_hash = str((item.get("pdf") or {}).get("sha256") or "")
        counts[pdf_hash] = counts.get(pdf_hash, 0) + 1
        case_id = str((item.get("case") or {}).get("case_id") or "")
        if case_id in selected_cases:
            raise ValueError(f"同一案卷被确认了多个审批 PDF: {case_id}")
        selected_cases.add(case_id)
    duplicates = [pdf_hash for pdf_hash, count in counts.items() if count > 1]
    if duplicates:
        raise ValueError(
            "同一审批 PDF 只能确认一个案卷: " + "，".join(duplicates)
        )

    applications = {
        str(item.get("case_id") or ""): item
        for item in dataset.get("applications") or []
    }
    unmatched = {
        str(item.get("sha256") or ""): item
        for item in dataset.get("unmatched_files") or []
        if (
            item.get("role") == APPROVAL_PDF_ROLE
            and _direct_inbox_pdf(root.resolve(), item) is not None
        )
    }
    for item in confirmed:
        pdf_hash = str((item.get("pdf") or {}).get("sha256") or "")
        case_id = str((item.get("case") or {}).get("case_id") or "")
        if pdf_hash not in unmatched:
            raise ValueError(f"审批 PDF 已不在待匹配列表: {pdf_hash}")
        if case_id not in applications:
            raise ValueError(f"案卷已不存在: {case_id}")
        path = ensure_within(
            root / Path(str(unmatched[pdf_hash].get("path") or "")),
            root,
        )
        if not path.is_file() or sha256_file(path) != pdf_hash:
            raise RuntimeError(f"审批 PDF 文件不存在或内容已变化: {path}")

    decided_at = _now()
    decision_history = review.setdefault("decisions", [])
    existing_history = {
        (
            str(item.get("review_id") or ""),
            str(item.get("decision") or ""),
        )
        for item in decision_history
    }
    for item in decisions:
        pdf = item.get("pdf") or {}
        case = item.get("case") or {}
        history_item = {
            "review_id": item.get("review_id", ""),
            "decision": item.get("decision", ""),
            "review_note": item.get("review_note", ""),
            "pdf_file_name": pdf.get("file_name", ""),
            "pdf_sha256": pdf.get("sha256", ""),
            "case_name": case.get("case_name", ""),
            "case_id": case.get("case_id", ""),
            "decided_at": decided_at,
        }
        history_key = (
            str(history_item["review_id"]),
            str(history_item["decision"]),
        )
        if history_key not in existing_history:
            decision_history.append(history_item)
            existing_history.add(history_key)

    applied_hashes: set[str] = set()
    for item in confirmed:
        pdf_hash = str((item.get("pdf") or {}).get("sha256") or "")
        case_id = str((item.get("case") or {}).get("case_id") or "")
        archive_reviewed_approval_pdf(
            dataset,
            root,
            unmatched[pdf_hash],
            applications[case_id],
            str(item.get("review_id") or ""),
            str(item.get("review_note") or ""),
        )
        applied_hashes.add(pdf_hash)

    excluded_ids = {
        str(item.get("review_id") or "")
        for item in decisions
        if item.get("decision") == "排除"
    }
    review["pending_reviews"] = [
        item
        for item in review.get("pending_reviews") or []
        if (
            str((item.get("pdf") or {}).get("sha256") or "")
            not in applied_hashes
            and str(item.get("review_id") or "") not in excluded_ids
        )
    ]
    review["last_applied_at"] = decided_at
    return len(confirmed)
