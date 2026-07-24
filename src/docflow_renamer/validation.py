from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .constants import EXCEL_FILE_NAME
from .file_utils import ensure_within, sha256_file


EXPECTED_SHEETS = [
    "申请汇总",
    "待补材料",
    "待审批PDF",
    "已完成",
    "本次变更",
    "说明",
]


def _all_file_records(application: dict[str, Any]) -> Iterable[dict[str, Any]]:
    for files in (application.get("materials") or {}).values():
        yield from files or []
    yield from (application.get("approval") or {}).get("pdfs") or []


def validate_dataset(
    data: dict[str, Any],
    root: Path,
    verify_hashes: bool = True,
) -> dict[str, Any]:
    root = root.resolve()
    errors: list[str] = []
    warnings: list[str] = []
    case_ids: set[str] = set()
    file_ids: set[str] = set()
    file_count = 0
    statuses: Counter[str] = Counter()

    for application in data.get("applications") or []:
        case_id = str(application.get("case_id") or "")
        if not case_id:
            errors.append("存在缺少 case_id 的案卷")
        elif case_id in case_ids:
            errors.append(f"重复 case_id: {case_id}")
        case_ids.add(case_id)
        statuses[str(application.get("status") or "unknown")] += 1

        relative_case_dir = str(application.get("case_directory") or "")
        try:
            case_dir = ensure_within(root / Path(relative_case_dir), root)
        except ValueError:
            errors.append(f"案卷路径越界: {relative_case_dir}")
            continue
        if not case_dir.is_dir():
            errors.append(f"案卷目录不存在: {relative_case_dir}")

        for item in _all_file_records(application):
            file_count += 1
            file_id = str(item.get("file_id") or "")
            if not file_id:
                errors.append(f"文件记录缺少 file_id: {item.get('path')}")
            elif file_id in file_ids:
                errors.append(f"重复 file_id: {file_id}")
            file_ids.add(file_id)
            _validate_file_record(
                item,
                root,
                errors,
                warnings,
                verify_hashes,
            )

    for item in data.get("unmatched_files") or []:
        file_count += 1
        _validate_file_record(
            item,
            root,
            errors,
            warnings,
            verify_hashes,
        )

    return {
        "valid": not errors,
        "applications": len(data.get("applications") or []),
        "files": file_count,
        "statuses": dict(statuses),
        "errors": errors,
        "warnings": warnings,
    }


def _validate_file_record(
    item: dict[str, Any],
    root: Path,
    errors: list[str],
    warnings: list[str],
    verify_hashes: bool,
) -> None:
    relative_path = str(item.get("path") or "")
    try:
        path = ensure_within(root / Path(relative_path), root)
    except ValueError:
        errors.append(f"文件路径越界: {relative_path}")
        return
    if not path.is_file():
        errors.append(f"文件不存在: {relative_path}")
        return
    expected_size = item.get("size")
    if isinstance(expected_size, int) and path.stat().st_size != expected_size:
        errors.append(f"文件大小不一致: {relative_path}")
    if verify_hashes and item.get("sha256"):
        if sha256_file(path) != item["sha256"]:
            errors.append(f"文件哈希不一致: {relative_path}")
    elif not item.get("sha256"):
        warnings.append(f"文件缺少 SHA-256: {relative_path}")


def validate_excel(root: Path, expected_applications: int) -> dict[str, Any]:
    excel_path = root.resolve() / EXCEL_FILE_NAME
    if not excel_path.is_file():
        return {
            "valid": False,
            "path": str(excel_path),
            "errors": ["Excel 文件不存在"],
        }
    workbook = load_workbook(excel_path, read_only=False, data_only=False)
    errors: list[str] = []
    try:
        if workbook.sheetnames != EXPECTED_SHEETS:
            errors.append(
                f"工作表不符合预期: {workbook.sheetnames}"
            )
        if "申请汇总" in workbook.sheetnames:
            summary = workbook["申请汇总"]
            actual_rows = max(0, summary.max_row - 1)
            if actual_rows != expected_applications:
                errors.append(
                    f"申请汇总行数错误: {actual_rows}，预期 {expected_applications}"
                )
            hyperlink_columns = (10, 11, 12, 13, 14, 16, 17)
            for row_index in range(2, summary.max_row + 1):
                for column_index in hyperlink_columns:
                    hyperlink = summary.cell(
                        row=row_index,
                        column=column_index,
                    ).hyperlink
                    if not hyperlink or not hyperlink.target:
                        continue
                    if not Path(hyperlink.target).exists():
                        errors.append(
                            "Excel 超链接目标不存在: "
                            f"{summary.cell(row=1, column=column_index).value}"
                            f" 第 {row_index} 行 -> {hyperlink.target}"
                        )
    finally:
        workbook.close()
    return {
        "valid": not errors,
        "path": str(excel_path),
        "sheets": EXPECTED_SHEETS,
        "errors": errors,
    }
