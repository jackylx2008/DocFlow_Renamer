from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .constants import EXCEL_FILE_NAME


STATUS_LABELS = {
    "materials_incomplete": "材料待补充",
    "materials_ready": "材料齐全，待审批PDF",
    "approval_pdf_unmatched": "审批PDF待确认",
    "approved": "审批完成",
    "needs_review": "待人工确认",
}
ROLE_LABELS = {
    "word": "Word申请单",
    "signed_application": "手签申请单",
    "worker_list": "施工人员名单",
    "safety_agreement": "安全生产及消防安全协议",
    "confined_space": "有限空间申请",
    "high_altitude": "高处作业申请",
    "special_work": "专项作业材料",
    "approval_pdf": "审批PDF",
}
SUMMARY_HEADERS = [
    "案卷状态",
    "材料完整性",
    "项目名称",
    "施工区域",
    "施工开始时间",
    "施工结束时间",
    "施工内容",
    "危险作业",
    "缺少材料",
    "Word申请单",
    "手签申请单",
    "施工人员名单",
    "专项作业材料",
    "安全协议",
    "审批编号",
    "审批PDF",
    "案卷目录",
    "案卷ID",
]


def _files(application: dict[str, Any], role: str) -> list[dict[str, Any]]:
    return list((application.get("materials") or {}).get(role) or [])


def _display_files(files: Iterable[dict[str, Any]]) -> str:
    return "；".join(str(item.get("current_name") or "") for item in files)


def _first_path(files: Iterable[dict[str, Any]], root: Path) -> str:
    for item in files:
        relative = str(item.get("path") or "")
        if relative:
            return str((root / Path(relative)).resolve())
    return ""


def _to_excel_date(value: Any) -> date | str:
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        return str(value or "")


def _summary_row(
    application: dict[str, Any], root: Path
) -> tuple[list[Any], dict[int, str]]:
    business = application.get("application") or {}
    approval = application.get("approval") or {}
    word_files = _files(application, "word")
    signed_files = _files(application, "signed_application")
    worker_files = _files(application, "worker_list")
    safety_files = _files(application, "safety_agreement")
    special_files = [
        *_files(application, "confined_space"),
        *_files(application, "high_altitude"),
        *_files(application, "special_work"),
    ]
    approval_files = list(approval.get("pdfs") or [])
    missing = [
        ROLE_LABELS.get(role, role)
        for role in application.get("missing_material_types") or []
    ]
    row = [
        STATUS_LABELS.get(application.get("status"), application.get("status")),
        "完整" if not missing else "缺少材料",
        business.get("项目名称", ""),
        business.get("施工区域", ""),
        _to_excel_date(business.get("施工开始时间")),
        _to_excel_date(business.get("施工结束时间")),
        business.get("施工内容", ""),
        business.get("危险作业", ""),
        "；".join(missing),
        _display_files(word_files),
        _display_files(signed_files),
        _display_files(worker_files),
        _display_files(special_files),
        _display_files(safety_files),
        approval.get("application_no", ""),
        _display_files(approval_files),
        application.get("case_directory", ""),
        application.get("case_id", ""),
    ]
    hyperlinks = {
        10: _first_path(word_files, root),
        11: _first_path(signed_files, root),
        12: _first_path(worker_files, root),
        13: _first_path(special_files, root),
        14: _first_path(safety_files, root),
        16: _first_path(approval_files, root),
        17: str((root / Path(application.get("case_directory") or "")).resolve()),
    }
    return row, hyperlinks


def _add_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    table_name: str,
    hyperlinks: list[dict[int, str]] | None = None,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    if not rows:
        sheet.append([""] * len(headers))

    header_fill = PatternFill("solid", fgColor="44546A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if hyperlinks:
        for row_index, links in enumerate(hyperlinks, start=2):
            for column_index, target in links.items():
                if not target:
                    continue
                cell = sheet.cell(row=row_index, column=column_index)
                cell.hyperlink = target
                cell.font = Font(color="0563C1", underline="single")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    )
    table = Table(
        displayName=table_name,
        ref=f"A1:{get_column_letter(len(headers))}{max(2, sheet.max_row)}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    widths = {
        "A": 20,
        "B": 12,
        "C": 18,
        "D": 22,
        "E": 13,
        "F": 13,
        "G": 28,
        "H": 22,
        "I": 30,
        "J": 38,
        "K": 38,
        "L": 38,
        "M": 38,
        "N": 38,
        "O": 16,
        "P": 38,
        "Q": 42,
        "R": 38,
    }
    for column, width in widths.items():
        if sheet.max_column >= ord(column) - ord("A") + 1:
            sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for column_index in (5, 6):
        for cell in sheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=sheet.max_row,
        ):
            cell[0].number_format = "yyyy-mm-dd"
    sheet.row_dimensions[1].height = 26


def _build_summary_sheets(
    workbook: Workbook, data: dict[str, Any], root: Path
) -> None:
    applications = list(data.get("applications") or [])
    all_rows: list[list[Any]] = []
    all_links: list[dict[int, str]] = []
    for application in applications:
        row, links = _summary_row(application, root)
        all_rows.append(row)
        all_links.append(links)
    _add_table_sheet(
        workbook,
        "申请汇总",
        SUMMARY_HEADERS,
        all_rows,
        "ApplicationsTable",
        all_links,
    )

    views = [
        (
            "待补材料",
            "MissingMaterialsTable",
            lambda item: bool(item.get("missing_material_types")),
        ),
        (
            "待审批PDF",
            "PendingApprovalTable",
            lambda item: item.get("status") == "materials_ready",
        ),
        (
            "已完成",
            "ApprovedTable",
            lambda item: item.get("status") == "approved",
        ),
    ]
    for title, table_name, predicate in views:
        rows: list[list[Any]] = []
        links: list[dict[int, str]] = []
        for application in applications:
            if predicate(application):
                row, row_links = _summary_row(application, root)
                rows.append(row)
                links.append(row_links)
        _add_table_sheet(
            workbook,
            title,
            SUMMARY_HEADERS,
            rows,
            table_name,
            links,
        )


def _build_unmatched_sheet(
    workbook: Workbook, data: dict[str, Any], root: Path
) -> None:
    headers = ["文件类型", "文件名", "相对路径", "SHA-256", "文件链接"]
    rows: list[list[Any]] = []
    links: list[dict[int, str]] = []
    for item in data.get("unmatched_files") or []:
        relative = str(item.get("path") or "")
        rows.append(
            [
                ROLE_LABELS.get(item.get("role"), item.get("role")),
                item.get("current_name", ""),
                relative,
                item.get("sha256", ""),
                "打开文件",
            ]
        )
        links.append({5: str((root / Path(relative)).resolve())})
    _add_table_sheet(
        workbook,
        "审批PDF待确认",
        headers,
        rows,
        "UnmatchedFilesTable",
        links,
    )


def _build_changes_sheet(workbook: Workbook, data: dict[str, Any]) -> None:
    headers = ["动作", "材料类型", "源路径", "目标路径", "结果", "案卷ID"]
    rows = [
        [
            item.get("action", ""),
            ROLE_LABELS.get(item.get("role"), item.get("role", "")),
            item.get("source", ""),
            item.get("target", ""),
            item.get("result", ""),
            item.get("case_id", ""),
        ]
        for item in data.get("changes") or []
    ]
    _add_table_sheet(
        workbook,
        "本次变更",
        headers,
        rows,
        "ChangesTable",
    )


def _build_notes_sheet(workbook: Workbook, data: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("说明")
    sheet.sheet_view.showGridLines = False
    applications = list(data.get("applications") or [])
    rows = [
        ["项目", "内容"],
        ["数据文件", "质保作业申请数据.json"],
        ["数据版本", data.get("dataset_revision", 0)],
        ["Schema版本", data.get("schema_version", 0)],
        ["生成时间", datetime.now().astimezone().isoformat(timespec="seconds")],
        ["案卷总数", len(applications)],
        [
            "审批完成",
            sum(item.get("status") == "approved" for item in applications),
        ],
        [
            "待审批PDF",
            sum(item.get("status") == "materials_ready" for item in applications),
        ],
        [
            "材料待补充",
            sum(bool(item.get("missing_material_types")) for item in applications),
        ],
        ["待确认文件", len(data.get("unmatched_files") or [])],
        [
            "使用说明",
            "本工作簿只用于人工调阅，唯一事实数据源是同目录下的质保作业申请数据.json。",
        ],
    ]
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="44546A")
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 80
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def export_excel(data: dict[str, Any], root: Path) -> Path:
    root = root.resolve()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_summary_sheets(workbook, data, root)
    _build_changes_sheet(workbook, data)
    _build_notes_sheet(workbook, data)

    output_path = root / EXCEL_FILE_NAME
    temporary_path = root / f".{EXCEL_FILE_NAME}.tmp.xlsx"
    workbook.save(temporary_path)
    try:
        os.replace(temporary_path, output_path)
    except PermissionError as exc:
        temporary_path.unlink(missing_ok=True)
        raise PermissionError(
            f"Excel 正在使用，JSON 已保留但未更新人工审查表: {output_path}"
        ) from exc
    return output_path
